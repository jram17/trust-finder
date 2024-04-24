import requests
from bs4 import BeautifulSoup
import re
import openpyxl



excel=openpyxl.Workbook()
print(excel.sheetnames)
sheet=excel.active
sheet.title='trustInformation'
print(excel.sheetnames)
sheet.append(["Trust Name","Trust Id"])


def getData(url):
    r = requests.get(url)
    soup = BeautifulSoup(r.text, 'html.parser')   
    return soup

def getNextPage(soup):
    page = soup.find('ul', {'class': 'pagination'})
    if page:
        current_active = page.find('li', {'class': 'active'})
        if current_active:
            next_sibling = current_active.find_next_sibling('li')
            if next_sibling:
                next_link = next_sibling.find('a')
                if next_link:
                    return next_link.get('href')
    return None

def removeParanthesis(name_element):
    if name_element:
        onclick_value = name_element.get('onclick')
        if onclick_value:
            info_in_parentheses = re.search(r'\("([^"]+)"\)', onclick_value).group(1)
            return info_in_parentheses


mainSource=requests.get('https://ngodarpan.gov.in/index.php/home/statewise')
soup=BeautifulSoup(mainSource.text,'html.parser')


stateTrusts=soup.find('ol',class_='rounded-list').find_all('li')

for stateTrust in stateTrusts:
    state=stateTrust.find('a').get('href')
    url=state
    while True:
        source=getData(url)
        trusts=source.find('tbody').find_all('tr')
        for trust in trusts:
            td_element=trust.find_all('td')
            sno=td_element[0].text
            name=td_element[1].text.strip()
            regNumber=td_element[2].text.split(',')[0]

            name_element = td_element[1].a
            voNumber=removeParanthesis(name_element)
        
            print(name,voNumber)
            sheet.append([name,voNumber])
        # print("Next Page URL:", getNextPage(source))
        url = getNextPage(source)
        if not url:
            break
    

excel.save('trustInfoIndia.xlsx')


